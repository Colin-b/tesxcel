from typing import Union
import openpyxl
from io import BytesIO


def assert_excel_content(
    file_path_or_actual_content: Union[str, bytes], expected_file_path: str
):
    if isinstance(file_path_or_actual_content, str):
        actual_workbook = openpyxl.open(file_path_or_actual_content)
    else:
        actual_workbook = openpyxl.load_workbook(BytesIO(file_path_or_actual_content))
    expected_workbook = openpyxl.open(expected_file_path)
    assert sorted(actual_workbook.sheetnames) == sorted(
        expected_workbook.sheetnames
    ), "Different sheet names"

    for sheet_name in actual_workbook.sheetnames:
        _assert_sheet_content(
            sheet_name,
            actual_workbook.get_sheet_by_name(sheet_name),
            expected_workbook.get_sheet_by_name(sheet_name),
        )


def _assert_sheet_content(sheet_name: str, actual_worksheet, expected_worksheet):
    assert (
        actual_worksheet.max_row == expected_worksheet.max_row
    ), f"Different number of rows in {sheet_name} sheet"
    assert (
        actual_worksheet.max_column == expected_worksheet.max_column
    ), f"Different number of columns in {sheet_name} sheet"

    for row_index, actual_row in enumerate(actual_worksheet.rows):
        expected_row = expected_worksheet[row_index + 1]
        for cell_index, actual_cell in enumerate(actual_row):
            expected_cell = expected_row[cell_index]
            assert (
                actual_cell.data_type == expected_cell.data_type
            ), f"Different cell type in row {row_index}, col {cell_index} in {sheet_name} sheet"
            assert (
                actual_cell.value == expected_cell.value
            ), f"Different cell content in row {row_index}, col {cell_index} in {sheet_name} sheet"
